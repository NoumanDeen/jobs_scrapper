import undetected_chromedriver as uc
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.chrome.options import Options
import time
import logging
import os
import json
import traceback
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import random
import requests
from fake_useragent import UserAgent

# Configure logging
logging.basicConfig(
    level=logging.INFO, 
    format='%(asctime)s - %(levelname)s: %(message)s',
    handlers=[
        logging.FileHandler('hiredly_scraper.log'),
        logging.StreamHandler()
    ]
)

class HireldyScraper:
    def __init__(self, chromedriver_path):
        """
        Initialize Hiredly Scraper
        
        :param chromedriver_path: Path to ChromeDriver executable
        """
        try:
            # Validate ChromeDriver path
            if not os.path.exists(chromedriver_path):
                raise FileNotFoundError(f"ChromeDriver not found at {chromedriver_path}")
            
            # Job titles and their corresponding URLs
            self.job_searches = [
                {
                    'title': 'ui/ux designer', 
                    'url': 'https://my.hiredly.com/jobs?search=ui%2Fux%20designer&state=Kuala%20Lumpur&state=Selangor&state=Putrajaya&state=Penang&state=Johor&state=Perlis&state=Kedah&state=Kelantan&state=Terengganu&state=Malacca&state=Negeri%20Sembilan&state=Pahang&state=Perak&state=Sabah&state=Sarawak&state=Labuan&state=Singapore&state=Overseas&job-type=5&job-type=4&job-type=3&job-type=2'
                },
                {
                    'title': 'ux designer', 
                    'url': 'https://my.hiredly.com/jobs?search=ux%20designer&state=Kuala+Lumpur&state=Selangor&state=Putrajaya&state=Penang&state=Johor&state=Perlis&state=Kedah&state=Kelantan&state=Terengganu&state=Malacca&state=Negeri+Sembilan&state=Pahang&state=Perak&state=Sabah&state=Sarawak&state=Labuan&state=Singapore&state=Overseas&job-type=5&job-type=4&job-type=3&job-type=2'
                },
                {
                    'title': 'ux design lead', 
                    'url': 'https://my.hiredly.com/jobs?search=ux%20design%20lead&state=Kuala+Lumpur&state=Selangor&state=Putrajaya&state=Penang&state=Johor&state=Perlis&state=Kedah&state=Kelantan&state=Terengganu&state=Malacca&state=Negeri+Sembilan&state=Pahang&state=Perak&state=Sabah&state=Sarawak&state=Labuan&state=Singapore&state=Overseas&job-type=5&job-type=4&job-type=3&job-type=2'
                },
                {
                    'title': 'ux design manager', 
                    'url': 'https://my.hiredly.com/jobs?search=ux%20design%20manager&state=Kuala+Lumpur&state=Selangor&state=Putrajaya&state=Penang&state=Johor&state=Perlis&state=Kedah&state=Kelantan&state=Terengganu&state=Malacca&state=Negeri+Sembilan&state=Pahang&state=Perak&state=Sabah&state=Sarawak&state=Labuan&state=Singapore&state=Overseas&job-type=5&job-type=4&job-type=3&job-type=2'
                },
                {
                    'title': 'product design lead', 
                    'url': 'https://my.hiredly.com/jobs?search=product%20design%20lead&state=Kuala+Lumpur&state=Selangor&state=Putrajaya&state=Penang&state=Johor&state=Perlis&state=Kedah&state=Kelantan&state=Terengganu&state=Malacca&state=Negeri+Sembilan&state=Pahang&state=Perak&state=Sabah&state=Sarawak&state=Labuan&state=Singapore&state=Overseas&job-type=5&job-type=4&job-type=3&job-type=2'
                },
                {
                    'title': 'product design manager', 
                    'url': 'https://my.hiredly.com/jobs?search=product%20design%20manager&state=Kuala+Lumpur&state=Selangor&state=Putrajaya&state=Penang&state=Johor&state=Perlis&state=Kedah&state=Kelantan&state=Terengganu&state=Malacca&state=Negeri+Sembilan&state=Pahang&state=Perak&state=Sabah&state=Sarawak&state=Labuan&state=Singapore&state=Overseas&job-type=5&job-type=4&job-type=3&job-type=2'
                }
            ]
            
            # Setup Chrome options with advanced configurations
            chrome_options = uc.ChromeOptions()
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--disable-extensions")
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_argument('--disable-web-security')
            chrome_options.add_argument('--allow-running-insecure-content')
            chrome_options.add_argument('--ignore-certificate-errors')
            
            # Use fake user agent
            ua = UserAgent()
            user_agent = ua.random
            chrome_options.add_argument(f'user-agent={user_agent}')
            
            # Randomize window size to appear more human-like
            window_sizes = [
                (1366, 768),
                (1920, 1080),
                (1600, 900),
                (1440, 900)
            ]
            size = random.choice(window_sizes)
            chrome_options.add_argument(f'--window-size={size[0]},{size[1]}')
            
            # Set ChromeDriver service
            service = Service(chromedriver_path)
            
            # Initialize the driver with advanced undetected mode
            logging.info("Initializing ChromeDriver for Hiredly...")
            self.driver = uc.Chrome(
                service=service, 
                options=chrome_options,
                enable_cdp_events=True
            )
            
            # Advanced page load settings
            self.driver.set_page_load_timeout(45)
            self.driver.implicitly_wait(10)
            
            logging.info("ChromeDriver initialized successfully")
            
            # Prepare output directory
            self.output_dir = 'hiredly_output'
            os.makedirs(self.output_dir, exist_ok=True)
            self.jobs = []
        
        except Exception as e:
            logging.error(f"Hiredly Scraper Initialization Error: {e}")
            logging.error(traceback.format_exc())
            raise

    def scrape_jobs(self):
        for search in self.job_searches:
            try:
                logging.info(f"Searching for jobs: {search['title']}")
                
                # Advanced navigation with retry mechanism
                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        self.driver.get(search['url'])
                        break
                    except Exception as nav_e:
                        if attempt == max_retries - 1:
                            raise
                        logging.warning(f"Navigation attempt {attempt + 1} failed: {nav_e}")
                        time.sleep(random.uniform(2, 5))
                
                # Advanced page load wait with multiple techniques
                time.sleep(random.uniform(7, 12))  # Random wait
                
                # Scroll to simulate human interaction
                self.driver.execute_script("window.scrollBy(0, window.innerHeight);")
                time.sleep(random.uniform(1, 3))
                self.driver.execute_script("window.scrollBy(0, -window.innerHeight);")
                
                # Take screenshot for debugging
                screenshot_path = os.path.join(self.output_dir, f"{search['title']}_page.png")
                self.driver.save_screenshot(screenshot_path)
                logging.info(f"Screenshot saved to {screenshot_path}")
                
                # Check if page loaded correctly
                current_url = self.driver.current_url
                logging.info(f"Current URL: {current_url}")
                
                # Detailed page source logging
                page_source = self.driver.page_source
                logging.info(f"Page source length: {len(page_source)} characters")
                
                # Check for potential blocking or captcha
                if "captcha" in page_source.lower() or "robot" in page_source.lower():
                    logging.error("Potential CAPTCHA or bot detection detected!")
                
                # Try multiple selectors with more detailed logging
                job_card_selectors = [
                    "div[data-testid='job-card']",
                    "div.job-card",
                    "div.MuiPaper-root",  # Fallback Material UI selector
                    "div[class*='job-card']",
                    "div[data-job-id]",
                    "div[class*='JobCard']",
                    "div[class*='job-listing']",
                    "div[data-cy='job-card']"
                ]
                
                job_cards = None
                for selector in job_card_selectors:
                    try:
                        # Wait for elements with explicit wait
                        job_cards = WebDriverWait(self.driver, 10).until(
                            EC.presence_of_all_elements_located((By.CSS_SELECTOR, selector))
                        )
                        logging.info(f"Selector '{selector}' found {len(job_cards)} elements")
                        if job_cards:
                            break
                    except Exception as sel_e:
                        logging.warning(f"Selector {selector} failed: {sel_e}")
                
                if not job_cards:
                    logging.error("No job cards found. Saving detailed page source for investigation.")
                    page_source_path = os.path.join(self.output_dir, f"{search['title']}_page_source.html")
                    with open(page_source_path, 'w', encoding='utf-8') as f:
                        f.write(page_source)
                    
                    # Additional debugging: print out body content
                    try:
                        body_text = self.driver.find_element(By.TAG_NAME, 'body').text
                        logging.error(f"Body text (first 1000 chars): {body_text[:1000]}")
                    except Exception as body_e:
                        logging.error(f"Could not extract body text: {body_e}")
                    
                    continue
                
                logging.info(f"Found {len(job_cards)} job cards")
                
                # Extract job details
                for card in job_cards:
                    try:
                        # Try multiple ways to extract job details
                        job_url = self.extract_job_url(card)
                        title = self.extract_text(card, ["h3[data-testid='job-card-title']", "h3.job-title"])
                        company = self.extract_text(card, ["p[data-testid='job-card-company']", "p.company-name"])
                        location = self.extract_text(card, ["p[data-testid='job-card-location']", "p.job-location"])
                        salary = self.extract_text(card, ["p[data-testid='job-card-salary']", "p.job-salary"], default="Not specified")
                        job_type = self.extract_text(card, ["p[data-testid='job-card-type']", "p.job-type"], default="Not specified")
                        posted_date = self.extract_text(card, ["p[data-testid='job-card-posted-date']", "p.job-posted-date"], default="Not specified")
                        
                        job_data = {
                            'Platform': 'Hiredly',
                            'Job Title Searched': search['title'],
                            'Job Title': title,
                            'Company': company,
                            'Location': location,
                            'Salary Range': salary,
                            'Job Type': job_type,
                            'Posted Date': posted_date,
                            'Job URL': job_url
                        }
                        
                        self.jobs.append(job_data)
                    
                    except Exception as card_e:
                        logging.error(f"Error processing job card: {card_e}")
                
                # Pagination handling
                self.handle_pagination(search['title'])
                
            except Exception as search_e:
                logging.error(f"Error during job search for {search['title']}: {search_e}")
                logging.error(traceback.format_exc())
        
        return self.jobs

    def extract_job_url(self, card):
        url_selectors = [
            "a[data-testid='job-card-link']",
            "a.job-card-link",
            "a[href*='/jobs/']"
        ]
        
        for selector in url_selectors:
            try:
                job_link = card.find_element(By.CSS_SELECTOR, selector)
                return job_link.get_attribute('href')
            except Exception:
                continue
        
        return "URL not found"

    def extract_text(self, card, selectors, default="Not specified"):
        for selector in selectors:
            try:
                element = card.find_element(By.CSS_SELECTOR, selector)
                return element.text.strip()
            except Exception:
                continue
        
        return default

    def handle_pagination(self, search_title):
        try:
            # Try multiple pagination selectors
            next_page_selectors = [
                "button[aria-label='Next page']",
                "button.next-page",
                "a[aria-label='Next']"
            ]
            
            max_pages = 5  # Limit to prevent infinite scraping
            current_page = 1
            
            while current_page < max_pages:
                try:
                    next_button = None
                    for selector in next_page_selectors:
                        try:
                            next_button = self.driver.find_element(By.CSS_SELECTOR, selector)
                            if next_button and next_button.is_enabled():
                                break
                        except Exception:
                            continue
                    
                    if not next_button or not next_button.is_enabled():
                        logging.info(f"No more pages for {search_title}")
                        break
                    
                    next_button.click()
                    current_page += 1
                    
                    # Wait for page to load
                    time.sleep(3)
                    
                    # Rescan job cards
                    job_cards = self.driver.find_elements(By.CSS_SELECTOR, "div[data-testid='job-card']")
                    
                    # Extract jobs from this page
                    for card in job_cards:
                        try:
                            job_url = self.extract_job_url(card)
                            title = self.extract_text(card, ["h3[data-testid='job-card-title']", "h3.job-title"])
                            company = self.extract_text(card, ["p[data-testid='job-card-company']", "p.company-name"])
                            
                            job_data = {
                                'Platform': 'Hiredly',
                                'Job Title Searched': search_title,
                                'Job Title': title,
                                'Company': company,
                                'Job URL': job_url
                            }
                            
                            self.jobs.append(job_data)
                        
                        except Exception as card_e:
                            logging.error(f"Error processing job card on page {current_page}: {card_e}")
                
                except Exception as page_e:
                    logging.error(f"Error navigating to page {current_page}: {page_e}")
                    break
        
        except Exception as pagination_e:
            logging.error(f"Pagination error for {search_title}: {pagination_e}")

    def save_results_excel(self, jobs):
        """
        Save scraped job listings to Excel
        
        :param jobs: List of job dictionaries
        """
        try:
            # Create timestamp for unique filename
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(self.output_dir, f'hiredly_jobs_{timestamp}.xlsx')
            
            # Create a new workbook
            wb = openpyxl.Workbook(write_only=False)
            ws = wb.active
            ws.title = "Job Listings"
            
            # Define headers
            headers = [
                'Platform', 'Job Title Searched', 'Job Title', 
                'Company', 'Location', 'Salary Range', 'Job Type', 
                'Posted Date', 'Job URL'
            ]
            
            # Write headers with styling
            header_font = Font(bold=True)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
            
            # Write job data
            for row, job in enumerate(jobs, 2):
                ws.cell(row=row, column=1, value=job['Platform'])
                ws.cell(row=row, column=2, value=job['Job Title Searched'])
                ws.cell(row=row, column=3, value=job['Job Title'])
                ws.cell(row=row, column=4, value=job['Company'])
                ws.cell(row=row, column=5, value=job['Location'])
                ws.cell(row=row, column=6, value=job['Salary Range'])
                ws.cell(row=row, column=7, value=job['Job Type'])
                ws.cell(row=row, column=8, value=job['Posted Date'])
                ws.cell(row=row, column=9, value=job['Job URL'])
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Save the workbook
            wb.save(output_file)
            
            logging.info(f"Total jobs found: {len(jobs)}")
            logging.info(f"Results saved to {output_file}")
            
            # Attempt to open the file
            try:
                os.startfile(output_file)
            except Exception as open_e:
                logging.warning(f"Could not automatically open file: {open_e}")
            
            return output_file
        
        except Exception as e:
            logging.error(f"Error saving results to Excel: {e}")
            logging.error(traceback.format_exc())
            return None

    def __del__(self):
        """Close browser on object destruction"""
        try:
            if hasattr(self, 'driver'):
                # Add explicit wait before quitting
                self.driver.close()
                time.sleep(1)
                self.driver.quit()
                logging.info("Hiredly Browser closed successfully")
        except Exception as e:
            logging.error(f"Error closing browser: {e}")
            logging.error(traceback.format_exc())

def main():
    # Potential ChromeDriver paths
    CHROMEDRIVER_PATHS = [
        r'C:\Users\numan\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe',
        r'D:\chromedriver\chromedriver.exe',
        r'C:\WebDrivers\chromedriver.exe'
    ]
    
    # Find the first existing ChromeDriver path
    CHROMEDRIVER_PATH = next((path for path in CHROMEDRIVER_PATHS if os.path.exists(path)), None)
    
    if not CHROMEDRIVER_PATH:
        print("Error: ChromeDriver not found. Please download and specify the correct path.")
        return
    
    scraper = None
    try:
        # Initialize and run Hiredly scraper
        scraper = HireldyScraper(CHROMEDRIVER_PATH)
        jobs = scraper.scrape_jobs()
        excel_file = scraper.save_results_excel(jobs)
        
        if excel_file:
            print(f"Jobs saved to {excel_file}")
    
    except Exception as e:
        logging.error(f"Hiredly Scraping Error: {e}")
        logging.error(traceback.format_exc())
        print(f"An error occurred: {e}")
    
    finally:
        # Ensure driver is closed even if an exception occurs
        if scraper and hasattr(scraper, 'driver'):
            try:
                scraper.driver.quit()
            except Exception as close_err:
                logging.error(f"Error in final driver cleanup: {close_err}")

if __name__ == "__main__":
    main()

