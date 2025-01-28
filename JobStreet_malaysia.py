import os
import time
import logging
import traceback
import json
import random
import urllib.parse
from selenium.webdriver.common.keys import Keys

import undetected_chromedriver as uc
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class JobStreetScraper:
    def __init__(self, chromedriver_path):
        """
        Initialize JobStreet Scraper
        
        :param chromedriver_path: Path to ChromeDriver executable
        """
        # Configure logging
        logging.basicConfig(
            level=logging.INFO, 
            format='%(asctime)s - %(levelname)s: %(message)s',
            handlers=[
                logging.FileHandler('jobstreet_scraper.log'),
                logging.StreamHandler()
            ]
        )
        
        try:
            # Validate ChromeDriver path
            if not os.path.exists(chromedriver_path):
                raise FileNotFoundError(f"ChromeDriver not found at {chromedriver_path}")
            
            # Set up Chrome options
            chrome_options = uc.ChromeOptions()
            chrome_options.add_argument('--start-maximized')
            chrome_options.add_argument('--disable-extensions')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--no-sandbox')
            
            # Initialize WebDriver
            self.driver = uc.Chrome(
                driver_executable_path=chromedriver_path, 
                options=chrome_options
            )
            
            # Implicit wait
            self.driver.implicitly_wait(10)
            
            # Job titles and their corresponding URLs
            self.job_searches = [
                {
                    'title': 'ux design manager', 
                    'url': 'https://my.jobstreet.com/ux-design-manager-jobs/in-Malaysia'
                },
                {
                    'title': 'product design lead', 
                    'url': 'https://my.jobstreet.com/product-design-lead-jobs/in-Malaysia'
                },
                {
                    'title': 'product design manager', 
                    'url': 'https://my.jobstreet.com/product-design-manager-jobs/in-Malaysia'
                }
            ]
        
        except Exception as e:
            logging.error(f"Initialization error: {e}")
            raise

    def scroll_and_wait(self, driver):
        """
        Quick scroll to load content
        """
        try:
            driver.execute_script("""
                var scrollHeight = document.body.scrollHeight;
                window.scrollTo(0, scrollHeight);
            """)
            time.sleep(0.5)  # Minimal wait
        except Exception as e:
            logging.warning(f"Quick scroll error: {e}")

    def scrape_jobstreet_jobs_with_pagination(self, driver, search_keyword, max_pages=None):
        """
        Scrape job listings with dynamic pagination support
        
        :param driver: Selenium WebDriver instance
        :param search_keyword: Keyword used for search
        :param max_pages: Optional maximum number of pages to scrape (None means all pages)
        :return: List of all job listings
        """
        all_jobs = []
        current_page = 1
        
        while True:
            try:
                # Quick page load wait
                time.sleep(1)
                
                # Quick scroll
                self.scroll_and_wait(driver)
                
                # Find job cards directly
                job_cards = driver.find_elements(By.CSS_SELECTOR, 'div.snwpn00[data-search-sol-meta]')
                
                if not job_cards:
                    logging.info("No more job cards found")
                    break
                
                # Scrape current page jobs
                jobs = []
                for card in job_cards:
                    try:
                        # Job Title
                        job_title_elem = card.find_element(By.CSS_SELECTOR, 'a[data-automation="jobTitle"]')
                        job_title = job_title_elem.text.strip()
                        job_url = job_title_elem.get_attribute('href')
                        
                        # Location
                        location_elem = card.find_element(By.CSS_SELECTOR, 'a[data-automation="jobLocation"]')
                        location = location_elem.text.strip()
                        
                        # Salary
                        try:
                            salary_elem = card.find_element(By.CSS_SELECTOR, 'span[data-automation="jobSalary"]')
                            salary = salary_elem.text.strip()
                        except:
                            salary = 'Not specified'
                        
                        # Posted Date
                        posted_date_elem = card.find_element(By.CSS_SELECTOR, 'span[data-automation="jobListingDate"]')
                        posted_date = posted_date_elem.text.strip()
                        
                        job_entry = {
                            'Platform': 'JobStreet',
                            'Job Title': job_title,
                            'Location': location,
                            'URL': job_url,
                            'Search Keyword': search_keyword,
                            'Salary': salary,
                            'Posted Date': posted_date
                        }
                        
                        jobs.append(job_entry)
                    
                    except Exception as e:
                        logging.warning(f"Error extracting job: {e}")
                
                all_jobs.extend(jobs)
                logging.info(f"Page {current_page}: Extracted {len(jobs)} jobs")
                
                # Find next page button
                try:
                    next_page_button = driver.find_element(By.CSS_SELECTOR, 'a[aria-label="Next"]')
                    
                    # Check if next page button is disabled or if we've reached max pages
                    if 'disabled' in next_page_button.get_attribute('class'):
                        logging.info("Reached last page")
                        break
                    
                    # Check if we've hit the optional max pages limit
                    if max_pages is not None and current_page >= max_pages:
                        logging.info(f"Reached maximum specified pages: {max_pages}")
                        break
                    
                    next_page_button.click()
                    current_page += 1
                    time.sleep(0.5)  # Minimal wait
                
                except:
                    logging.info("No more pages to navigate")
                    break
            
            except Exception as e:
                logging.error(f"Page {current_page} error: {e}")
                break
        
        logging.info(f"Total jobs scraped: {len(all_jobs)} across {current_page} pages")
        return all_jobs

    def scrape_jobs(self):
        """
        Scrape job listings from JobStreet for multiple job titles
        """
        all_jobs = []
        
        for job_search in self.job_searches:
            job_title = job_search['title']
            search_url = job_search['url']
            
            try:
                logging.info(f"Scraping: {job_title}")
                self.driver.get(search_url)
                time.sleep(1)  # Minimal page load wait
                
                # Remove max_pages parameter to scrape all pages
                jobs = self.scrape_jobstreet_jobs_with_pagination(self.driver, job_title)
                all_jobs.extend(jobs)
                
                logging.info(f"Found {len(jobs)} jobs for {job_title}")
            
            except Exception as e:
                logging.error(f"Scraping error for {job_title}: {e}")
        
        return all_jobs

    def save_results_excel(self, jobs):
        """
        Save job results to Excel with specified columns
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Listings"
        
        # Simplified headers
        headers = [
            'Platform', 'Job Title', 'Location', 
            'URL', 'Search Keyword', 'Salary', 
            'Posted Date'
        ]
        
        # Header styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Populate data
        for row, job in enumerate(jobs, 2):
            ws.cell(row=row, column=1, value=job.get('Platform', 'N/A'))
            ws.cell(row=row, column=2, value=job.get('Job Title', 'N/A'))
            ws.cell(row=row, column=3, value=job.get('Location', 'N/A'))
            ws.cell(row=row, column=4, value=job.get('URL', 'N/A'))
            ws.cell(row=row, column=5, value=job.get('Search Keyword', 'N/A'))
            ws.cell(row=row, column=6, value=job.get('Salary', 'N/A'))
            ws.cell(row=row, column=7, value=job.get('Posted Date', 'N/A'))
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save the workbook
        excel_filename = f'jobstreet_jobs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(excel_filename)
        logging.info(f"Job results saved to {excel_filename}")
        
        return excel_filename

    def close(self):
        """
        Close the browser driver
        """
        try:
            if self.driver:
                self.driver.quit()
        except Exception as e:
            logging.error(f"Error closing driver: {e}")

def main():
    try:
        # Path to ChromeDriver (update this to your actual path)
        chromedriver_path = r"C:\Users\numan\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe"
        
        # Initialize scraper
        scraper = JobStreetScraper(chromedriver_path)
        
        try:
            # Scrape jobs
            jobs = scraper.scrape_jobs()
            
            # Save results to Excel
            if jobs:
                scraper.save_results_excel(jobs)
            else:
                logging.warning("No jobs found.")
        
        finally:
            # Ensure driver is closed
            scraper.close()
    
    except Exception as e:
        logging.error(f"Scraping failed: {e}")
        logging.error(traceback.format_exc())

if __name__ == "__main__":
    main()
